import openpyxl, sys
path = r'icpep-schedule-tracker-1.0.0.xlsx'
wb = openpyxl.load_workbook(path)
out = []
out.append('Sheet names: ' + str(wb.sheetnames))

for sname in wb.sheetnames:
    ws = wb[sname]
    out.append(f'\n--- {sname} (rows={ws.max_row}, cols={ws.max_column}) ---')
    count = 0
    for row in ws.iter_rows(values_only=True):
        if any(v is not None for v in row):
            out.append(str(row[:12]))
            count += 1
        if count >= 20:
            out.append('...(truncated)')
            break

with open('excel_inspect_out.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(out))
print('Done')
